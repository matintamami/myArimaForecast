from pandas import read_csv
from pandas import read_excel
from pandas import datetime
from pandas import DataFrame
from statsmodels.tsa.arima_model import ARIMA
from matplotlib import pyplot
from sklearn.metrics import mean_squared_error
from math import sqrt
import openpyxl

# def parser(x):
#     return datetime.strptime('190'+x, '%Y-%m')

exceldata = openpyxl.load_workbook('E:/Matin/Work Programming/Bentoel/bentoel-dataset(fixed).xlsx')
ws = exceldata.get_sheet_by_name("Sheet1")
data_dict = []
for row in ws.iter_rows('H{}:H{}'.format(4,ws.max_row)):
    for cell in row:
        if cell.value != None:
            data_dict.append(cell.value)
print data_dict
# series = read_excel('E:/Matin/Work Programming/Bentoel/bentoel-dataset(fixed).xlsx',header=None,usecols=[7],squeeze=True)
x = data_dict
size = int(len(x) * 0.31)
print size
train, test = x[0:size], x[size:len(x)]
history = [x for x in train]
predictions = list()
selisih = list()
for t in range(len(test)):
    model = ARIMA(history, order=(0, 1, 0))
    model_fit = model.fit(disp=0)
    output = model_fit.forecast()
    yhat = output[0]
    predictions.append(yhat)
    obs = test[t]
    history.append(obs)
    selisih.append(yhat - obs)
    print('predicted=%f, expected=%f' % (yhat, obs))
error = mean_squared_error(test, predictions)
rmse = sqrt(mean_squared_error(test,predictions))
se = (sum(selisih)) / len(data_dict)
print('Test MSE: %.3f' % error)
print('Test RSME: %.3f' % rmse)
print('Test SE: %.3f' % se)
# plot
pyplot.plot(test)
pyplot.plot(predictions, color='red')
pyplot.show()